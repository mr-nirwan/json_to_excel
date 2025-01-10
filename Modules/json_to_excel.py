import streamlit as st
import pandas as pd
import fitz  # PyMuPDF
import io
import re
import json
import zipfile
import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import logging
import traceback
import base64

# --------------------------------------------------------------
# Highlight utility functions
# --------------------------------------------------------------
def normalize_text(text: str) -> str:
    """
    Normalize text by:
    - Replacing different dash characters (–, —) with a single hyphen (-)
    - Collapsing multiple whitespace into a single space
    - Stripping leading/trailing spaces
    """
    text = text.replace("–", "-")
    text = text.replace("—", "-")
    text = re.sub(r"\s+", " ", text)
    return text.strip()

def highlight_snippet_with_chunks(page: fitz.Page, 
                                  snippet: str, 
                                  min_words: int = 3, 
                                  min_chars: int = 15,
                                  color: tuple = (1, 1, 0)) -> None:
    """
    Attempt to highlight 'snippet' on the given 'page' using a chunk-based approach.
    1) First try to find and highlight the entire snippet.
    2) If not found, break the snippet into chunks from left to right.
    """
    snippet = normalize_text(snippet)
    if not snippet:
        return

    # Try exact match first
    found = page.search_for(snippet)
    if found:
        for inst in found:
            highlight = page.add_highlight_annot(inst)
            if highlight:
                highlight.set_colors({"stroke": color})
                highlight.update()
        return

    # Chunk-based fallback
    words = snippet.split()
    start_idx = 0
    while start_idx < len(words):
        chunk_found = False
        # Try progressively smaller chunks from the end
        for end_idx in range(len(words), start_idx, -1):
            candidate_words = words[start_idx:end_idx]
            candidate_str = " ".join(candidate_words)

            if len(candidate_words) < min_words or len(candidate_str) < min_chars:
                continue

            found = page.search_for(candidate_str)
            if found:
                for inst in found:
                    highlight = page.add_highlight_annot(inst)
                    if highlight:
                        highlight.set_colors({"stroke": color})
                        highlight.update()

                start_idx = end_idx
                chunk_found = True
                break

        if not chunk_found:
            start_idx += 1

def highlight_text_in_pdf(file_bytes: bytes, 
                          texts_to_highlight: list, 
                          highlight_color: tuple = (1, 1, 0)) -> bytes:
    """
    Highlights occurrences of each string in `texts_to_highlight`
    within the PDF given by `file_bytes`.
    Returns the highlighted PDF as bytes.
    """
    doc = fitz.open(stream=file_bytes, filetype="pdf")

    for page in doc:
        for snippet in texts_to_highlight:
            snippet_str = str(snippet).strip()
            if not snippet_str:
                continue
            highlight_snippet_with_chunks(
                page, 
                snippet_str, 
                min_words=2,       # you can adjust these
                min_chars=8,       # you can adjust these
                color=highlight_color
            )

    output_buffer = io.BytesIO()
    doc.save(output_buffer)
    doc.close()
    output_buffer.seek(0)
    return output_buffer.read()

def save_df_to_excel(df_processed):
    try:
        def clean_text_for_excel(text):
            """
            Remove illegal XML characters and potential formula strings from a string before inserting into Excel.
            Escapes any text that starts with characters potentially interpreted by Excel as a formula.
            """
            if isinstance(text, (str, float, int)):  # Now handling numbers as well
                text = str(text)  # Convert to string to clean
                # Escape strings that start with =, +, -, or @
                if text.startswith(('=', '+', '-', '@')):
                    text = f"'{text}"
                # Remove illegal XML characters
                text = re.sub(r"[^\x09\x0A\x0D\x20-\x7E\x85\xA0-\uD7FF\uE000-\uFDCF\uFDE0-\uFFFD]", "", text)
            return text
        if df_processed.empty:
            raise ValueError("The DataFrame is empty. Please provide a DataFrame with data.")
            
        # Your base logic to save df to Excel without specific formatting
        workbook = Workbook()
        worksheet = workbook.active
        # Styling
        header_fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Write the DataFrame headers and apply styles
        for col_num, column_title in enumerate(df_processed.columns, start=1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = clean_text_for_excel(column_title)
            cell.fill = header_fill
            cell.font = header_font

        # Write DataFrame data with enhanced error logging and highlighting errors in red
        for row_num, row in enumerate(df_processed.itertuples(index=False), start=2):
            for col_num, value in enumerate(row, start=1):
                try:
                    worksheet.cell(row=row_num, column=col_num, value=value)
                except Exception as e:
                    try:
                        # Attempt to clean the text and retry setting the cell value
                        cleaned_value = clean_text_for_excel(value)
                        worksheet.cell(row=row_num, column=col_num, value=cleaned_value)
                    except Exception as e:
                        # Log the error with detailed information and the full stack trace
                        logging.warning(f"Error setting value '{value}' in row {row_num}, column={col_num} after cleaning: {traceback.format_exc()}")
                        # Here, you might want to set a placeholder value or apply special formatting
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = f"{str(value)[:20]}..."  # Truncated placeholder
                        # Optionally apply a style, like red_fill, to highlight the error  
                        #cell.fill = red_fill  # Use the red_fill style defined in your function       
        # Set row height and apply alignment
        for row in worksheet.iter_rows():
            worksheet.row_dimensions[row[0].row].height = 15  # Set the height of the row
            for cell in row:
                cell.alignment = alignment
        # Enable filters on the header row
        worksheet.auto_filter.ref = worksheet.dimensions
        # Save the workbook to BytesIO object
        output_data = BytesIO()
        workbook.save(output_data)
        output_data.seek(0)
        # Convert to base64 for download
        b64 = base64.b64encode(output_data.read()).decode()

        return b64
    except Exception as e:
        logging.error(f"Error occurred while saving DataFrame to Excel: {e}")
        return None, None


# --------------------------------------------------------------
# Streamlit app
# --------------------------------------------------------------
def main():
    st.title("PDF Highlighter and Excel Generator")

    # Step 1) Upload PDF
    uploaded_pdf = st.file_uploader("Upload a PDF", type=["pdf"])

    # Step 2) Decide how many JSON inputs (default 14). 
    # Let the user specify the number of JSON text areas.
    num_json_inputs = st.number_input("Number of JSON inputs", min_value=1, value=10, step=1)

    # Collect JSON text in a list
    json_texts = []
    for i in range(num_json_inputs):
        json_text = st.text_area(f"JSON Input #{i+1}", height=100, key=f"json_input_{i}")
        json_texts.append(json_text)

    # Button to process
    if st.button("Process JSON and Show DataFrame"):
        if not uploaded_pdf:
            st.error("No PDF file uploaded.")
            return

        # Initialize a list to hold all dataframes
        dfs = []
        # A set to collect all text for highlighting
        all_texts_for_highlight = set()

        # Parse each non-empty JSON
        valid_json_found = False
        for i, json_text in enumerate(json_texts):
            if not json_text.strip():
                continue  # skip empty

            valid_json_found = True
            try:
                data = json.loads(json_text)
            except Exception as e:
                st.error(f"Invalid JSON in input #{i+1}! Error: {e}")
                return

            # -- ADDED PART: Auto-convert if missing data_rows and top-level is a list --
            if "data_rows" not in data:
                if isinstance(data, list):
                    # Wrap the list into a dict with key "data_rows"
                    data = {"data_rows": data}
                else:
                    st.error(f"JSON in input #{i+1} does not contain 'data_rows' key and is not a top-level list.")
                    return
            # ---------------------------------------------------------------------------

            data_rows = data["data_rows"]
            if not isinstance(data_rows, list):
                st.error(f"`data_rows` in input #{i+1} should be a list. Invalid format.")
                return
            
            # Convert to DataFrame
            df_part = pd.DataFrame(data_rows)

            # Add text from these rows to our highlight set
            for row in data_rows:
                for val in row.values():
                    val_str = str(val).strip()
                    if val_str:
                        all_texts_for_highlight.add(val_str)

            # Optionally rename columns to avoid collisions 
            # if multiple JSONs have columns named the same.
            df_part.columns = [f"{col}_set{i+1}" for col in df_part.columns]

            dfs.append(df_part)

        if not valid_json_found:
            st.error("No valid JSON data provided in any text area.")
            return

        # Concatenate all dataframes column-wise
        if len(dfs) == 1:
            final_df = dfs[0]
        else:
            # Column-wise concatenation; adjust 'axis=1' for horizontal
            # If row lengths differ, you'll get NaNs for mismatched positions
            final_df = pd.concat(dfs, axis=1)

        st.success("DataFrame created successfully from all JSONs!")
        st.dataframe(final_df)
        pdf_name = uploaded_pdf.name  # e.g., "mydoc.pdf"
        base_name = os.path.splitext(pdf_name)[0]  # e.g., "mydoc"
        excel_filename = f"{base_name}.xlsx"
        zip_filename = f"{base_name}.zip"
        # Prepare an Excel file (instead of CSV, per your request)
        b64 = save_df_to_excel(final_df)

        # Convert base64 string to bytes
        excel_bytes = base64.b64decode(b64)

        # Step 3) Read PDF bytes
        pdf_bytes = uploaded_pdf.read()

        # Step 4) Highlight
        st.info("Highlighting text in PDF (this could take a moment)...")
        try:
            all_texts_for_highlight_list = list(all_texts_for_highlight)
            highlighted_pdf_bytes = highlight_text_in_pdf(pdf_bytes, all_texts_for_highlight_list)
            st.success("Highlighting completed!")

            # We'll create a zip file in memory that contains:
            #   1) The highlighted PDF
            #   2) The Excel file (same name as PDF but with .xlsx)
            # The zip will also have the same base name as the PDF.

            # Write into a zip archive in-memory
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                # Add the highlighted PDF
                zf.writestr(pdf_name, highlighted_pdf_bytes)
                # Add the Excel file
                zf.writestr(excel_filename, excel_bytes)
            zip_buffer.seek(0)

            # Download button for the zip
            st.download_button(
                label="Download ZIP (Highlighted PDF + Excel)",
                data=zip_buffer,
                file_name=zip_filename,
                mime="application/zip"
            )

        except Exception as e:
            st.error(f"Error during highlighting: {e}")

if __name__ == "__main__":
    main()
